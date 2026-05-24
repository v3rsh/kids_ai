"""
Генератор Excel-выгрузок «Безопасные рисунки».

Принципиальное правило: ``registry.xlsx`` **не хранится на диске** и
не пересобирается на каждое событие. Файл собирается из БД по запросу
`/export` и `/export_shortlist`, отдаётся в чат attachment'ом
и забывается. Сервис возвращает ``bytes``.

Источник правды по формату Excel — `docs/registry-spec.md`.
"""
from __future__ import annotations

import time
import uuid as uuid_pkg
from datetime import datetime, timezone
from io import BytesIO
from typing import Literal, Sequence
from zoneinfo import ZoneInfo

from loguru import logger
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import select
from sqlalchemy.orm import selectinload

from config import COMPETITION_YEAR
from database.db import get_session
from database.models import (
    Application,
    IntakeMode,
    JuryMember,
    JuryRound,
    JuryStatus,
    JuryVote,
    JuryVoteState,
    JuryVoteValue,
)
from services import pools as pools_service
from utils.contracts import PoolKey

MSK = ZoneInfo("Europe/Moscow")


# =====================================================================
# Транслитерация и заголовки колонок жюри
# =====================================================================
#
# Используется паспортный стандарт ICAO Doc 9303 (актуальная редакция
# Приказа МВД РФ № 889 / Приказа МИД РФ № 4271 от 2014).
# Регистр выхода — Title-case (Shcherbak, не SHCHERBAK).

_ICAO_9303_MAP: dict[str, str] = {
    "А": "A", "Б": "B", "В": "V", "Г": "G", "Д": "D",
    "Е": "E", "Ё": "E", "Ж": "ZH", "З": "Z", "И": "I",
    "Й": "I", "К": "K", "Л": "L", "М": "M", "Н": "N",
    "О": "O", "П": "P", "Р": "R", "С": "S", "Т": "T",
    "У": "U", "Ф": "F", "Х": "KH", "Ц": "TS", "Ч": "CH",
    "Ш": "SH", "Щ": "SHCH", "Ъ": "IE", "Ы": "Y", "Ь": "",
    "Э": "E", "Ю": "IU", "Я": "IA",
}


def transliterate_icao_9303(text: str) -> str:
    """Транслитерация кириллицы по ICAO Doc 9303.

    Прозрачно проходит через символы, не входящие в таблицу
    (латиница, цифры, дефисы и т. п.) — это нужно для fallback'а на
    смешанные ФИО (например, `O'Брайан`).
    """
    out: list[str] = []
    for ch in text:
        if ch.isupper():
            out.append(_ICAO_9303_MAP.get(ch, ch))
        elif ch.islower():
            mapped = _ICAO_9303_MAP.get(ch.upper(), ch)
            out.append(mapped.lower())
        else:
            out.append(ch)
    return "".join(out)


def jury_column_header(full_name: str, round_no: int) -> str:
    """Шапка динамической колонки листа `Голосование жюри`.

    Шаблон: ``<Фамилия>.<И>_r<N>`` (например, ``Vinokurova.E_r1``).

    Алгоритм разбора ``full_name``:
        1. Разбиваем по whitespace.
        2. Если ≥2 токенов и оба непустые: фамилия = первый токен в
           Title-case, инициал = первая буква второго токена в upper.
        3. Иначе fallback: ``<full_name>_r<N>`` (без транслитерации).
    """
    tokens = full_name.strip().split()
    if len(tokens) >= 2 and tokens[0] and tokens[1]:
        surname_t = transliterate_icao_9303(tokens[0])
        # Инициал ограничиваем РОВНО одной буквой латиницы: если первая
        # буква имени даёт многосимвольную транслитерацию (Ю→IU, Я→IA,
        # Ж→ZH, Х→KH, Ц→TS, Ч→CH, Ш→SH, Щ→SHCH, Ъ→IE) — берём только
        # первый символ результата, иначе фиксированная ширина колонки
        # 14 (см. docs/registry-spec.md) рассыпается на длинных инициалах.
        initial_t = transliterate_icao_9303(tokens[1][0])[:1]
        if surname_t and initial_t:
            return f"{surname_t.title()}.{initial_t.upper()}_r{round_no}"
    return f"{full_name.strip()}_r{round_no}"


# =====================================================================
# Публичные helpers (имена файлов выгрузок)
# =====================================================================


def registry_export_filename(
    kind: Literal["registry", "shortlist"],
    now_msk: datetime | None = None,
) -> str:
    """Имя файла on-demand выгрузки (см. ``docs/registry-spec.md``).

    Шаблон: ``{kind}_BR-{COMPETITION_YEAR}_{YYYY-MM-DD}_{HH-MM}.xlsx``.

    Примеры:
        >>> from datetime import datetime
        >>> registry_export_filename(
        ...     "registry",
        ...     now_msk=datetime(2026, 6, 15, 14, 32, tzinfo=MSK),
        ... )
        'registry_BR-2026_2026-06-15_14-32.xlsx'

    Аргументы:
        kind: тип выгрузки — ``registry`` (основной реестр) или
            ``shortlist`` (шорт-лист топ-10 по пулам).
        now_msk: момент вызова в ``Europe/Moscow``; если ``None`` —
            берётся ``datetime.now(MSK)``. Параметр явно вынесен наружу,
            чтобы тесты получали стабильные имена.

    Возвращает:
        Имя файла, готовое для передачи в pybotx attachment.

    Бросает:
        ``ValueError`` — если ``kind`` не входит в допустимый набор.
    """
    if kind not in ("registry", "shortlist"):
        raise ValueError(f"Unknown registry kind: {kind!r}")
    now = now_msk if now_msk is not None else datetime.now(MSK)
    return (
        f"{kind}_BR-{COMPETITION_YEAR}_"
        f"{now:%Y-%m-%d}_{now:%H-%M}.xlsx"
    )


# =====================================================================
# Helpers значений строк
# =====================================================================


def view_command_or_link(app: Application) -> str:
    """Значение поля «Команда/ссылка просмотра файлов».

    - ``IntakeMode.LINKS`` → ``app.cloud_link`` (URL папки участника)
      или пустая строка, если ссылка ещё не получена;
    - ``IntakeMode.FILES`` → ``/files <br_id>`` (текстовая команда
      модератора в чате).

    Та же функция переиспользуется в шорт-листе.
    """
    if app.intake_mode is IntakeMode.LINKS:
        return app.cloud_link or ""
    return f"/files {app.br_id}"


def contact_field(app: Application) -> str:
    """Значение поля «Контакт».

    Приоритет:
    1. ``parent_contact`` — то, что родитель явно ввёл на шаге
       «Контакт» в анкете (email или телефон);
    2. ``@<parent_ad_login>`` — корпоративный логин из CTS, если есть;
    3. ``HUID: <uuid>`` — последний fallback (HUID всегда доступен).
    """
    if getattr(app, "parent_contact", None):
        return app.parent_contact
    if app.parent_ad_login:
        return f"@{app.parent_ad_login}"
    return f"HUID: {app.parent_huid}"


def jury_outcome(app: Application) -> str:
    """Значение поля «Итог по жюри».

    Производное от ``Application.jury_status``:
    - ``не_передано_жюри`` → ``не оценивалась``;
    - ``в_топ-10`` → ``в топ-10``;
    - ``не_вошло_в_топ-10`` → ``не вошло в топ-10``;
    - ``на_голосовании`` → пусто (пул ещё не завершён).
    """
    if app.jury_status is JuryStatus.NE_PEREDANO_ZHYURI:
        return "не оценивалась"
    if app.jury_status is JuryStatus.V_TOP_10:
        return "в топ-10"
    if app.jury_status is JuryStatus.NE_VOSHLO_V_TOP_10:
        return "не вошло в топ-10"
    return ""


# =====================================================================
# Стили и колонки
# =====================================================================

# Soft-порог производительности — при превышении пишем WARNING в лог.
_DURATION_WARN_MS = 5000

# Стили шапки и разделителей.
_HEADER_FONT = Font(bold=True)
_HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9D9D9")
_GROUP_FONT = Font(bold=True)
_GROUP_FILL = PatternFill(fill_type="solid", fgColor="BFBFBF")
_EMPTY_GROUP_FONT = Font(italic=True)
_EMPTY_GROUP_FILL = PatternFill(fill_type="solid", fgColor="D9D9D9")
_WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")
_GROUP_ALIGN = Alignment(horizontal="left", vertical="center")

# Колонки основного листа `Реестр`: (заголовок, ширина, wrap_text).
# Порядок жёсткий, соответствует разделу 2.2 `docs/registry-spec.md`.
_MAIN_COLUMNS: list[tuple[str, int, bool]] = [
    ("ID заявки", 14, False),                                # 1
    ("Дата и время подачи (Europe/Moscow)", 22, False),      # 2
    ("ФИО родителя", 28, False),                             # 3
    ("Подразделение", 24, False),                            # 4
    ("Контакт", 18, False),                                  # 5
    ("Имя ребёнка", 16, False),                              # 6
    ("Возраст ребёнка", 8, False),                           # 7
    ("Возрастная категория", 12, False),                     # 8
    ("Трек", 22, False),                                     # 9
    ("Название работы", 28, False),                          # 10
    ("Описание работы", 40, True),                           # 11
    ("Количество файлов", 10, False),                        # 12
    ("Команда/ссылка просмотра файлов", 32, False),          # 13
    ("Статус модерации", 18, False),                         # 14
    ("Комментарий модератора", 30, True),                    # 15
    ("Статус жюри", 18, False),                              # 16
    ("Статус голосования", 18, False),                       # 17
    ("Номер для голосования", 12, False),                    # 18
    ("Потенциал для мерча", 18, False),                      # 19
    ("Возможный дубль", 10, False),                          # 20
    ("Связанная заявка", 14, False),                         # 21
    ("Актуальная версия заявки", 10, False),                 # 22
    ("Голосов «Достоин» в р.1", 12, False),                  # 23
    ("Голосов «Достоин» в р.2", 12, False),                  # 24
    ("Голосов «Достоин» в р.3", 12, False),                  # 25
    ("Итоговый раунд", 10, False),                           # 26
    ("Итог по жюри", 18, False),                             # 27
    ("Определено жребием", 12, False),                       # 28
    ("Позиция в пуле", 10, False),                           # 29
]

# Фиксированные колонки листа `Голосование жюри`.
# Динамические колонки приклеиваются справа в _build_jury_detail_sheet.
_JURY_FIXED_COLUMNS: list[tuple[str, int, bool]] = [
    ("ID заявки", 14, False),
    ("Трек", 22, False),
    ("Возрастная категория", 12, False),
]
_JURY_DYNAMIC_WIDTH = 14

# Колонки листа `Шорт-лист`: (заголовок, ширина, wrap_text).
# 13 колонок в зафиксированном порядке (см. `docs/registry-spec.md`).
_SHORTLIST_COLUMNS: list[tuple[str, int, bool]] = [
    ("ID заявки", 14, False),                                # 1
    ("ФИО родителя", 28, False),                             # 2
    ("Контакт", 18, False),                                  # 3
    ("Имя ребёнка", 16, False),                              # 4
    ("Возраст", 8, False),                                   # 5
    ("Возрастная категория", 12, False),                     # 6
    ("Трек", 22, False),                                     # 7
    ("Название работы", 28, False),                          # 8
    ("Описание работы", 40, True),                           # 9
    ("Команда/ссылка просмотра файлов", 32, False),          # 10
    ("Определено жребием", 12, False),                       # 11
    ("Позиция в пуле", 10, False),                           # 12
    ("Потенциал для мерча", 18, False),                      # 13
]

# Текст пустого пула на листе `Шорт-лист`.
# В шаблон подставляется TOP_N из конфига при первом вызове рендера.
_EMPTY_POOL_TEXT_TEMPLATE = "Нет работ в топ-{top_n} для этого пула"


# =====================================================================
# Низкоуровневые helpers форматирования и openpyxl
# =====================================================================


def _yesno_or_blank(value: bool) -> str:
    """`да` / пусто — для булевых колонок реестра."""
    return "да" if value else ""


def _to_msk_iso(dt: datetime) -> str:
    """Дата подачи → ISO-строка `YYYY-MM-DD HH:MM` в Europe/Moscow.

    ``Application.created_at`` хранится в БД как naive UTC
    (``default=datetime.utcnow``). Здесь явно присваиваем UTC-tz,
    переводим в Europe/Moscow и форматируем.
    """
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(MSK).strftime("%Y-%m-%d %H:%M")


def _apply_columns_header(
    ws: Worksheet, columns: Sequence[tuple[str, int, bool]]
) -> None:
    """Заполнить шапку (строка 1) и установить ширины колонок."""
    for idx, (title, width, _wrap) in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
        cell = ws.cell(row=1, column=idx, value=title)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL


def _apply_wrap_text(
    ws: Worksheet, row: int, columns: Sequence[tuple[str, int, bool]]
) -> None:
    """Включить wrap_text для нужных колонок строки `row`."""
    for idx, (_title, _w, wrap) in enumerate(columns, start=1):
        if wrap:
            ws.cell(row=row, column=idx).alignment = _WRAP_ALIGN


def _set_freeze_and_filter(
    ws: Worksheet, freeze_at: str, n_cols: int, n_rows: int
) -> None:
    """Заморозка шапки и autofilter на весь диапазон."""
    ws.freeze_panes = freeze_at
    if n_cols >= 1 and n_rows >= 1:
        last_letter = get_column_letter(n_cols)
        ws.auto_filter.ref = f"A1:{last_letter}{n_rows}"


# =====================================================================
# Лист `Реестр`
# =====================================================================


def _row_for_main_sheet(app: Application) -> list:
    """29 значений одной строки основного листа."""
    n_files = len(app.files) if app.files else 0
    return [
        app.br_id,                                                 # 1
        _to_msk_iso(app.created_at),                               # 2
        app.parent_full_name,                                      # 3
        app.parent_division,                                       # 4
        contact_field(app),                                        # 5
        app.child_name,                                            # 6
        app.child_age,                                             # 7
        app.age_category.value,                                    # 8
        app.track.value,                                           # 9
        app.title,                                                 # 10
        app.description,                                           # 11
        n_files,                                                   # 12
        view_command_or_link(app),                                 # 13
        app.moderation_status.value,                               # 14
        app.moderator_comment or "",                               # 15
        app.jury_status.value,                                     # 16
        app.voting_status.value,                                   # 17
        "",                                                        # 18 (поле дизайнера)
        app.merch_potential or "",                                 # 19
        _yesno_or_blank(app.is_possible_duplicate),                # 20
        app.related_application_br_id or "",                       # 21
        "да" if app.is_actual_version else "нет",                  # 22
        app.jury_round1_yes,                                       # 23
        app.jury_round2_yes,                                       # 24
        app.jury_round3_yes,                                       # 25
        app.jury_final_round if app.jury_final_round is not None else "",  # 26
        jury_outcome(app),                                         # 27
        _yesno_or_blank(app.jury_decided_by_lot),                  # 28
        app.pool_position if app.pool_position is not None else "",        # 29
    ]


def _build_main_sheet(
    ws: Worksheet, applications: Sequence[Application]
) -> tuple[int, int]:
    """Заполнить лист `Реестр`. Возвращает (n_rows, n_cols).

    Сортировка применяется на этапе SQL (`ORDER BY br_id ASC`).
    """
    _apply_columns_header(ws, _MAIN_COLUMNS)
    for row_offset, app in enumerate(applications, start=2):
        for col_idx, value in enumerate(_row_for_main_sheet(app), start=1):
            ws.cell(row=row_offset, column=col_idx, value=value)
        _apply_wrap_text(ws, row_offset, _MAIN_COLUMNS)
    n_cols = len(_MAIN_COLUMNS)
    n_rows = 1 + len(applications)
    _set_freeze_and_filter(ws, "A2", n_cols, n_rows)
    return n_rows, n_cols


# =====================================================================
# Лист `Голосование жюри`
# =====================================================================


def _build_jury_detail_sheet(
    ws: Worksheet,
    applications: Sequence[Application],
    votes: Sequence[JuryVote],
    rounds_by_id: dict[uuid_pkg.UUID, JuryRound],
    jury_by_huid: dict[uuid_pkg.UUID, JuryMember],
) -> tuple[int, int]:
    """Заполнить лист `Голосование жюри`. Возвращает (n_rows, n_cols).

    Фильтр строк — только заявки, у которых есть хотя бы одна запись
    `JuryVote` любого состояния. Сортировка строк — `br_id ASC`.

    Динамические колонки — каждая (`JuryMember`, `round_no`) по
    фактически проведённым раундам, сортированы по полному имени судьи
    и номеру раунда.

    Формат значения — числовой 1/0/пусто. Учитываются только голоса
    в состоянии ``SUBMITTED``; ``DRAFT`` отображается как пустая ячейка.
    """
    # Оси: какие заявки и какие судьи × раунды имеют хотя бы один голос.
    app_ids_with_votes = {v.application_id for v in votes}
    apps_with_votes = sorted(
        (a for a in applications if a.id in app_ids_with_votes),
        key=lambda a: a.br_id,
    )

    jury_huids_in_votes = {v.jury_huid for v in votes}
    jury_members_sorted = sorted(
        (jury_by_huid[h] for h in jury_huids_in_votes if h in jury_by_huid),
        key=lambda j: j.full_name,
    )

    round_ids_in_votes = {v.round_id for v in votes}
    round_nos_sorted = sorted({
        rounds_by_id[r].round_no
        for r in round_ids_in_votes
        if r in rounds_by_id
    })

    # Шапка: фикс + динамика.
    columns: list[tuple[str, int, bool]] = list(_JURY_FIXED_COLUMNS)
    # Маппинг (jury_huid, round_no) → 1-based column_index.
    dyn_col: dict[tuple[uuid_pkg.UUID, int], int] = {}
    for jury in jury_members_sorted:
        for r_no in round_nos_sorted:
            columns.append(
                (jury_column_header(jury.full_name, r_no),
                 _JURY_DYNAMIC_WIDTH, False)
            )
            dyn_col[(jury.huid, r_no)] = len(columns)

    _apply_columns_header(ws, columns)

    # Индекс голосов: (application_id, jury_huid, round_no) → JuryVote.
    vote_lookup: dict[tuple[uuid_pkg.UUID, uuid_pkg.UUID, int], JuryVote] = {}
    for v in votes:
        if v.round_id not in rounds_by_id:
            continue
        r_no = rounds_by_id[v.round_id].round_no
        vote_lookup[(v.application_id, v.jury_huid, r_no)] = v

    # Строки.
    for row_offset, app in enumerate(apps_with_votes, start=2):
        ws.cell(row=row_offset, column=1, value=app.br_id)
        ws.cell(row=row_offset, column=2, value=app.track.value)
        ws.cell(row=row_offset, column=3, value=app.age_category.value)
        for (huid, r_no), col_idx in dyn_col.items():
            v = vote_lookup.get((app.id, huid, r_no))
            if v is None or v.state is not JuryVoteState.SUBMITTED:
                continue  # пусто
            ws.cell(
                row=row_offset, column=col_idx,
                value=1 if v.vote is JuryVoteValue.YES else 0,
            )

    n_cols = len(columns)
    n_rows = 1 + len(apps_with_votes)
    # Freeze: шапка + 3 фикс. колонки → D2.
    _set_freeze_and_filter(ws, "D2", n_cols, n_rows)
    return n_rows, n_cols


# =====================================================================
# Производительность и лог
# =====================================================================


def _log_duration(
    kind: str, duration_ms: float, n_rows: int, n_cols: int,
) -> None:
    logger.info(
        "registry build done",
        kind=kind,
        rows=n_rows,
        cols=n_cols,
        duration_ms=round(duration_ms, 1),
    )
    if duration_ms > _DURATION_WARN_MS:
        logger.warning(
            "registry build exceeded soft threshold",
            kind=kind,
            duration_ms=round(duration_ms, 1),
            threshold_ms=_DURATION_WARN_MS,
            hint=(
                "см. docs/registry-spec.md — рассмотреть кэш TTL=60s"
            ),
        )


# =====================================================================
# Pure-сборка workbook (тестируется без БД)
# =====================================================================


def _render_registry_workbook(
    applications: Sequence[Application],
    votes: Sequence[JuryVote],
    rounds_by_id: dict[uuid_pkg.UUID, JuryRound],
    jury_by_huid: dict[uuid_pkg.UUID, JuryMember],
) -> tuple[bytes, int, int]:
    """Собрать `registry.xlsx` в bytes; вернуть (bytes, total_cols, ...).

    Возвращает кортеж ``(payload, total_cols, total_rows)`` для лога.
    """
    wb = Workbook()
    main_ws = wb.active
    main_ws.title = "Реестр"
    n_rows_main, n_cols_main = _build_main_sheet(main_ws, applications)

    jury_ws = wb.create_sheet("Голосование жюри")
    _build_jury_detail_sheet(
        jury_ws, applications, votes, rounds_by_id, jury_by_huid,
    )

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue(), n_cols_main, n_rows_main


# =====================================================================
# SQL-getters (один запрос на сущность, никаких N+1)
# =====================================================================


async def _fetch_all_applications(session) -> list[Application]:
    """Все заявки + связанные файлы одним запросом."""
    stmt = (
        select(Application)
        .options(selectinload(Application.files))
        .order_by(Application.br_id.asc())
    )
    return list((await session.scalars(stmt)).all())


async def _fetch_jury_axes(
    session,
) -> tuple[
    list[JuryVote],
    dict[uuid_pkg.UUID, JuryRound],
    dict[uuid_pkg.UUID, JuryMember],
]:
    """Голоса + индексы раундов и судей для листа `Голосование жюри`.

    Три отдельных запроса (votes, rounds, jury) с последующей
    индексацией в памяти — для агрегации (application_id, jury_huid,
    round_no) в матрице ячеек. N+1 не возникает.
    """
    votes = list((await session.scalars(select(JuryVote))).all())
    rounds = list((await session.scalars(select(JuryRound))).all())
    jury = list((await session.scalars(select(JuryMember))).all())
    rounds_by_id = {r.id: r for r in rounds}
    jury_by_huid = {j.huid: j for j in jury}
    return votes, rounds_by_id, jury_by_huid


# =====================================================================
# Публичный API (контракт RegistryService)
# =====================================================================


async def build_registry_xlsx() -> bytes:
    """Собрать полный реестр заявок в XLSX-bytes.

    Структура:
        - лист ``Реестр`` — 29 колонок (поля заявки + агрегированные
          поля жюри);
        - лист ``Голосование жюри`` — детализация по голосам с
          динамическими колонками `Фамилия.И_rN`.

    Все данные собираются из БД одной транзакцией (`selectinload`
    для `Application.files`, отдельные запросы для голосов/раундов/
    судей). На диск ничего не пишется — возвращаются непосредственно
    ``bytes``.
    """
    t0 = time.perf_counter()
    logger.info("registry build start", kind="registry")

    async with get_session()() as session:
        applications = await _fetch_all_applications(session)
        votes, rounds_by_id, jury_by_huid = await _fetch_jury_axes(session)

    payload, n_cols, n_rows = _render_registry_workbook(
        applications=applications,
        votes=votes,
        rounds_by_id=rounds_by_id,
        jury_by_huid=jury_by_huid,
    )

    duration_ms = (time.perf_counter() - t0) * 1000
    _log_duration("registry", duration_ms, n_rows=n_rows, n_cols=n_cols)
    return payload


# =====================================================================
# Лист `Шорт-лист`
# =====================================================================


def _row_for_shortlist(app: Application) -> list:
    """13 значений одной строки шорт-листа."""
    return [
        app.br_id,                                                 # 1
        app.parent_full_name,                                      # 2
        contact_field(app),                                        # 3
        app.child_name,                                            # 4
        app.child_age,                                             # 5
        app.age_category.value,                                    # 6
        app.track.value,                                           # 7
        app.title,                                                 # 8
        app.description,                                           # 9
        view_command_or_link(app),                                 # 10
        _yesno_or_blank(app.jury_decided_by_lot),                  # 11
        app.pool_position if app.pool_position is not None else "",  # 12
        app.merch_potential or "",                                 # 13
    ]


def _merge_row_with_fill(
    ws: Worksheet,
    row: int,
    n_cols: int,
    text: str,
    font: Font,
    fill: PatternFill,
) -> None:
    """Заполнить строку-разделитель: merged через n_cols + заливка/шрифт."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = font
    cell.fill = fill
    cell.alignment = _GROUP_ALIGN
    # Заливка нужна на всех ячейках merge'а — иначе видно только в первой.
    for col in range(2, n_cols + 1):
        ws.cell(row=row, column=col).fill = fill


def _build_shortlist_sheet(
    ws: Worksheet,
    pool_list: Sequence[PoolKey],
    apps_by_pool: dict[tuple, list[Application]],
    top_n: int,
) -> tuple[int, int]:
    """Заполнить лист `Шорт-лист`. Возвращает (n_rows, n_cols).

    Группировка строк — по `pool_list` в порядке, который отдал
    `services.pools.all_pools()`. Сортировка внутри пула —
    `pool_position ASC, br_id ASC`. Пустой пул выводится со
    строкой-разделителем + строкой-подписью «Нет работ в топ-N».
    """
    _apply_columns_header(ws, _SHORTLIST_COLUMNS)
    n_cols = len(_SHORTLIST_COLUMNS)
    current_row = 2

    for pool in pool_list:
        # Строка-разделитель пула.
        header_text = f"{pool.track.value} / {pool.age_category.value}"
        _merge_row_with_fill(
            ws, current_row, n_cols, header_text, _GROUP_FONT, _GROUP_FILL,
        )
        current_row += 1

        apps = apps_by_pool.get((pool.track, pool.age_category), [])
        if not apps:
            _merge_row_with_fill(
                ws, current_row, n_cols,
                _EMPTY_POOL_TEXT_TEMPLATE.format(top_n=top_n),
                _EMPTY_GROUP_FONT, _EMPTY_GROUP_FILL,
            )
            current_row += 1
            continue

        # Стабильная сортировка по позиции в пуле.
        apps_sorted = sorted(
            apps,
            key=lambda a: (
                a.pool_position if a.pool_position is not None else 9999,
                a.br_id,
            ),
        )
        for app in apps_sorted:
            for col_idx, value in enumerate(_row_for_shortlist(app), start=1):
                ws.cell(row=current_row, column=col_idx, value=value)
            _apply_wrap_text(ws, current_row, _SHORTLIST_COLUMNS)
            current_row += 1

    n_rows = current_row - 1
    _set_freeze_and_filter(ws, "A2", n_cols, n_rows)
    return n_rows, n_cols


def _render_shortlist_workbook(
    pool_list: Sequence[PoolKey],
    apps_by_pool: dict[tuple, list[Application]],
    top_n: int,
) -> tuple[bytes, int, int]:
    """Собрать `shortlist.xlsx` в bytes; вернуть (bytes, n_cols, n_rows)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Шорт-лист"
    n_rows, n_cols = _build_shortlist_sheet(ws, pool_list, apps_by_pool, top_n)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue(), n_cols, n_rows


async def _fetch_top10_applications(session) -> list[Application]:
    """Только заявки со статусом жюри `в топ-10`."""
    stmt = (
        select(Application)
        .where(Application.jury_status == JuryStatus.V_TOP_10)
        .order_by(Application.br_id.asc())
    )
    return list((await session.scalars(stmt)).all())


async def build_shortlist_xlsx() -> bytes:
    """Собрать XLSX шорт-листа — топ-N по каждому пулу.

    Структура:
        - один лист ``Шорт-лист``;
        - 13 колонок (см. ``docs/registry-spec.md``);
        - строки сгруппированы по пулам в порядке
          ``services.pools.all_pools()``; число пулов не зашивается;
        - внутри пула — сортировка `pool_position ASC, br_id ASC`;
        - пустые пулы выводятся со строкой-подписью «Нет работ в топ-N».

    Не пишет на диск, возвращает ``bytes``.
    """
    from config import TOP_N  # локально, чтобы тестировалось без TOP_N

    t0 = time.perf_counter()
    logger.info("registry build start", kind="shortlist")

    async with get_session()() as session:
        top_apps = await _fetch_top10_applications(session)
        pool_list = pools_service.all_pools()

    apps_by_pool: dict[tuple, list[Application]] = {}
    for app in top_apps:
        key = (app.track, app.age_category)
        apps_by_pool.setdefault(key, []).append(app)

    payload, n_cols, n_rows = _render_shortlist_workbook(
        pool_list=pool_list, apps_by_pool=apps_by_pool, top_n=TOP_N,
    )

    duration_ms = (time.perf_counter() - t0) * 1000
    _log_duration("shortlist", duration_ms, n_rows=n_rows, n_cols=n_cols)
    return payload


__all__ = [
    "MSK",
    "registry_export_filename",
    "transliterate_icao_9303",
    "jury_column_header",
    "view_command_or_link",
    "contact_field",
    "jury_outcome",
    "build_registry_xlsx",
    "build_shortlist_xlsx",
]
